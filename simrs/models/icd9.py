# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook

class icd9(models.Model):
    _name = 'simrs.icd9'
    _description = 'ICD-9 Prosedur'

    icd9_code = fields.Char(string='Kode Prosedur', required=True)
    name = fields.Char(string='Nama Prosedur', required=True)
    file_data = fields.Binary(string="Upload File XLSX")

    _sql_constraints = [
        ('icd9_code_unique', 'unique(icd9_code)', 'Kode ICD-9 harus unik!')
    ]

    def import_xls(self):
        if not self.file_data:
            raise UserError("Silakan unggah file sebelum mengimpor.")

        file = base64.b64decode(self.file_data)
        workbook = load_workbook(filename=io.BytesIO(file), data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Lewati header
            icd9_code = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if row[1] else ''

            if not icd9_code or not name:
                continue

            existing = self.env['simrs.icd9'].search([('icd9_code', '=', icd9_code)], limit=1)
            if not existing:
                self.env['simrs.icd9'].create({
                    'icd9_code': icd9_code,
                    'name': name,
                })
