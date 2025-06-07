# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook

class ICD10(models.Model):
    _name = 'simrs.icd10'
    _description = 'ICD-10 Diagnosa'

    icd10_code = fields.Char(string='Kode ICD-10')
    name = fields.Char(string='Nama Diagnosa')
    file_data = fields.Binary(string="Upload File XLSX")  # Tempat upload file

    _sql_constraints = [
        ('icd10_code_unique', 'unique(icd10_code)', 'Kode ICD-10 harus unik!')
    ]

    def name_get(self):
        result = []
        for rec in self:
            name = f"{rec.icd10_code} - {rec.name}"
            result.append((rec.id, name))
        return result

    def import_xls(self):
        """Import ICD-10 data dari file XLSX berisi 2 kolom: Kode, Nama Diagnosa"""
        if not self.file_data:
            raise UserError("Silakan unggah file sebelum mengimpor.")

        file = base64.b64decode(self.file_data)
        workbook = load_workbook(filename=io.BytesIO(file), data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Baris pertama header
            icd10_code = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if row[1] else ''

            if not icd10_code or not name:
                continue  # Skip baris kosong

            # Cek apakah kode sudah ada
            existing = self.env['simrs.icd10'].search([('icd10_code', '=', icd10_code)], limit=1)
            if not existing:
                self.env['simrs.icd10'].create({
                    'icd10_code': icd10_code,
                    'name': name,
                })

    
